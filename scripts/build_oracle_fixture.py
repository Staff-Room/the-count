#!/usr/bin/env python3
"""
Regenerate the v1 income-statement conformance-oracle fixture.

The canonical source is the committed workbook:
    tests/fixtures/oracle/ZW_Consulting_TY2025_Schedule_C_Income_Statement_V1.xlsx

This script derives the diff-friendly, deterministic artifacts the test suite reads:
    - one CSV per worksheet (stable column order, ISO dates, empty cells = "")
    - income_statement_v1.json  (the parsed Schedule C P&L, the inner-loop oracle)
    - SOURCE_SHA256              (integrity pin for the workbook)

It is idempotent: running it on an unchanged workbook produces byte-identical
outputs. Re-run after replacing the workbook, then commit the regenerated files.

Usage:
    python scripts/build_oracle_fixture.py
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
from pathlib import Path

import openpyxl

ORACLE_DIR = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "oracle"
WORKBOOK = ORACLE_DIR / "ZW_Consulting_TY2025_Schedule_C_Income_Statement_V1.xlsx"

# Worksheet title -> output CSV filename.
SHEET_FILES = {
    "Schedule C P&L": "schedule_c_pl.csv",
    "Detail": "detail.csv",
    "Flags for Rozella": "flags_for_rozella.csv",
    "Methodology & Sources": "methodology_and_sources.csv",
}

# Schedule C P&L line labels that are totals/subtotals rather than coded buckets.
GROSS_INCOME_LINE = "Line 7"
TOTAL_EXPENSES_LINE = "Line 28"
NET_PROFIT_LINE = "Line 31"


def _cell(value: object) -> str:
    """Render a cell deterministically for CSV output."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        # All dates in this workbook are date-only.
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def _trim(row: list[object]) -> list[object]:
    out = list(row)
    while out and (out[-1] is None or out[-1] == ""):
        out.pop()
    return out


def sheet_to_csv(ws: "openpyxl.worksheet.worksheet.Worksheet", path: Path) -> int:
    rows = []
    width = 0
    for raw in ws.iter_rows(values_only=True):
        trimmed = _trim(list(raw))
        rows.append(trimmed)
        width = max(width, len(trimmed))
    # Drop fully-blank trailing rows.
    while rows and not rows[-1]:
        rows.pop()
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh, lineterminator="\n")
        for r in rows:
            padded = [_cell(v) for v in r] + [""] * (width - len(r))
            writer.writerow(padded)
    return len(rows)


def parse_income_statement(ws: "openpyxl.worksheet.worksheet.Worksheet") -> dict:
    """Parse the Schedule C P&L sheet into a machine-readable oracle."""
    header = [_cell(ws.cell(row=1, column=c).value) for c in range(1, 4)]
    subtitle = _cell(ws.cell(row=2, column=1).value)
    prepared = _cell(ws.cell(row=3, column=1).value)

    line_items: list[dict] = []
    section = None
    gross_income = total_expenses = net_profit = None

    for raw in ws.iter_rows(min_row=5, values_only=True):
        col_a = _cell(raw[0]) if len(raw) > 0 else ""
        if not col_a:
            continue
        if col_a in ("INCOME", "EXPENSES"):
            section = col_a
            continue
        if not col_a.startswith("Line "):
            continue

        description = raw[1] if len(raw) > 1 else None
        amount = raw[2] if len(raw) > 2 else None
        num_txns = raw[3] if len(raw) > 3 else None
        notes = raw[4] if len(raw) > 4 else None
        amount = float(amount) if isinstance(amount, (int, float)) else None

        item = {
            "line": col_a,
            "description": description,
            "amount": amount,
            "num_txns": int(num_txns) if isinstance(num_txns, (int, float)) else None,
            "section": section,
            "notes": notes,
        }
        line_items.append(item)

        if col_a == GROSS_INCOME_LINE:
            gross_income = amount
        elif col_a == TOTAL_EXPENSES_LINE:
            total_expenses = amount
        elif col_a == NET_PROFIT_LINE:
            net_profit = amount

    return {
        "title": header[0],
        "entity": subtitle,
        "prepared": prepared,
        "version": "V1",
        "basis": "cash",
        "tax_year": 2025,
        "line_items": line_items,
        "totals": {
            "gross_income": gross_income,
            "total_expenses": total_expenses,
            "net_profit_or_loss": net_profit,
        },
    }


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    digest = hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
    (ORACLE_DIR / "SOURCE_SHA256").write_text(
        f"{digest}  {WORKBOOK.name}\n", encoding="utf-8"
    )

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)

    for title, filename in SHEET_FILES.items():
        if title not in wb.sheetnames:
            raise SystemExit(f"Expected sheet missing from workbook: {title!r}")
        n = sheet_to_csv(wb[title], ORACLE_DIR / filename)
        print(f"wrote {filename:<32} ({n} rows)")

    income_statement = parse_income_statement(wb["Schedule C P&L"])
    (ORACLE_DIR / "income_statement_v1.json").write_text(
        json.dumps(income_statement, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print("wrote income_statement_v1.json")
    print(f"SOURCE_SHA256 = {digest}")
    t = income_statement["totals"]
    print(
        "oracle totals: gross_income={gross_income} "
        "total_expenses={total_expenses} "
        "net_profit_or_loss={net_profit_or_loss}".format(**t)
    )


if __name__ == "__main__":
    main()
